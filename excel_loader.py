from openpyxl import load_workbook
import pandas as pd
import io

def upload_and_extract_excel(uploaded_file):
    in_mem_file = io.BytesIO(uploaded_file.read())
    wb = load_workbook(in_mem_file, data_only=True)
    ws = wb.active
    table = []
    for row in ws.iter_rows():
        table.append([cell.value for cell in row])
    df_value = pd.DataFrame(table)
    return df_value
